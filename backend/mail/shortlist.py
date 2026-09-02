"""
Hiring-manager shortlists.

The hand-off at the end of the funnel. Once a role's assessments are graded,
the recruiter sends its best candidates to the manager who owns the seat: a
table in the email body they can read on a phone, and a spreadsheet attached
for anyone who would rather sort it themselves.

BY DEFAULT THE MANAGER DOES NOT GET THE SCORE. Not the number, not the band,
not the verdict, not the per-criterion marks -- see SHORTLIST_SHOW_SCORES in
config.py for why that is a policy rather than an oversight. Rank position is
included, because the order is the recommendation and hiding it would just make
the list arbitrary; the magnitude is not, because a "78" next to a name decides
the interview before the manager has read a word of the work.

The recruiter can override that per send: the Shortlist tab has an "AI score"
tick that puts the number in the SPREADSHEET -- the attachment and the download
are the same file, so both carry it -- and nowhere else. The email body stays a
list of names and links whichever way the tick sits, because the mail is what
gets read on a phone in ten seconds and is the surface where a number does the
most deciding for the least reading. `include_scores` carries that choice from
the click down to build_xlsx(); left as None it falls back to the config
default, so a caller that does not care does not have to decide.

THE TICK IS THE RECRUITING TEAM'S. A hiring-manager account cannot ask for it
-- server.py refuses the request rather than quietly answering without it, and
that is the check that actually enforces the policy, since nothing in this
module knows who is calling.

That makes RANK the only thing on this list a manager can weigh, which is why a
verdict the AI did not finish must never reach it. A part-filled grid is
renormalised to 100 by the scorer -- one row marked 5 renormalises to exactly
100.0 -- so it sorts above every honestly earned score, and there is no number
on this page for a reader to find odd. `rows()` draws from `top_candidates`,
which holds those verdicts out; `held_back()` hands them back to the recruiter
so nobody is dropped silently; and if the rule is switched off, every surface
here -- the table, the plain-text copy, the spreadsheet -- says "partly graded"
beside the name.

Everything here is pure apart from send_shortlist(): rows(), build_xlsx() and
build_email() take a role and return data, so the dashboard's preview shows
exactly what the send would deliver.
"""

import io
import logging
import re
from datetime import datetime

from backend.config import (
    BREVO_SENDER_EMAIL,
    BREVO_SENDER_NAME,
    DASHBOARD_BASE_URL,
    PUBLIC_BASE_URL,
    SHORTLIST_DASHBOARD_LINK,
    SHORTLIST_MAX,
    SHORTLIST_REQUIRE_COMPLETE_GRID,
    SHORTLIST_SHOW_SCORES,
    SHORTLIST_SIZE,
)
from backend.mail import brevo_client
from backend.mail.text import esc as _esc, first_name as _first_name
from backend.db import store
from backend.grading import rubric_pack
from backend.grading import tier_resolver

log = logging.getLogger(__name__)


class ShortlistError(RuntimeError):
    """A shortlist that cannot be built or sent, with a reason worth showing."""


# Ajaia's palette, matching the reminder email so the two read as one sender.
NAVY = "#001d6b"
INK = "#1b1c1c"
LINE = "#e4e7ec"
MUTED = "#5b6270"


# ---------------------------------------------------------------------------
# Rows
# ---------------------------------------------------------------------------

def _fmt_date(value) -> str:
    """
    "2026-08-10T17:13:18.338Z" -> "10 Aug 2026".

    The portal's CSV hands back an ISO string, Mongo hands back a datetime for
    anything we wrote ourselves, and a blank is a real possibility for a row
    that never submitted. All three land here.
    """
    if isinstance(value, datetime):
        return value.strftime("%d %b %Y")
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).strftime("%d %b %Y")
    except ValueError:
        return text[:10]


def role_view(role: dict, tier: str = "") -> dict:
    """
    The role as one tier of it: the posting's own name, and its own filename.

    A tiered assignment sends two shortlists, and they must not look like two
    copies of one. The manager gets "Senior AI Strategist -- top 10" and "AI
    Strategist -- top 10" as separate mails with separate attachments; without
    the slug suffix the second spreadsheet would overwrite the first in a
    downloads folder, which is a small thing that goes wrong at exactly the
    wrong moment.

    Returns the role unchanged when there is no tier, so every caller can pass
    one and none of them has to check first.
    """
    if not tier:
        return role
    labels = tier_resolver.posting_labels(role.get("slug"))
    return {
        **role,
        "title": labels.get(tier) or f"{role.get('title')} ({tier})",
        "slug": f"{role.get('slug') or 'role'}-{tier}",
    }


def scores_on(include_scores: bool | None = None) -> bool:
    """
    Whether this hand-off carries the AI score.

    One function so the rows, the spreadsheet's caption and the dashboard's
    preview cannot answer it three ways. None means "nobody chose", which is
    the config default -- an explicit False from a caller is a choice and stays
    a choice.
    """
    return SHORTLIST_SHOW_SCORES if include_scores is None else bool(include_scores)


def rows(job_id: int, limit: int = SHORTLIST_SIZE, tier: str = "",
         include_scores: bool | None = None) -> list[dict]:
    """
    The top `limit` candidates for a role as flat, sendable rows.

    Rank is assigned here rather than in the template, so the email, the
    spreadsheet and the dashboard preview cannot disagree about who is third.

    `tier` narrows to one posting where an assignment is marked at two. Rank
    only means something among people marked against the same anchors, so a
    tiered role's shortlist is two lists rather than one merged one -- a new
    graduate and a seven-year consultant ranked against each other on a
    40-point background row is not a recommendation anybody should act on.
    """
    size = max(1, min(int(limit or SHORTLIST_SIZE), SHORTLIST_MAX))
    default_tier = rubric_pack.default_tier_for_slug(
        (store.get_role(job_id) or {}).get("slug")) if tier else None
    out = []
    for i, sub in enumerate(store.top_candidates(job_id, size, tier=tier or None,
                                                 default_tier=default_tier),
                            start=1):
        row = {
            "rank": i,
            "submission_id": sub["_id"],
            "name": sub.get("candidate_name") or "(no name)",
            "email": sub.get("candidate_email") or "",
            "resume_link": sub.get("resume_link") or "",
            "video_link": sub.get("video_link") or "",
            # The assessment itself, on the portal, where the manager can read
            # what the candidate actually wrote.
            "assessment_url": sub.get("admin_url") or "",
            "submitted_at": _fmt_date(sub.get("submitted_at")),
            # Whether the AI finished this candidate's rubric. Normally nobody
            # here is provisional -- `top_candidates` holds them out -- but the
            # rule is a switch, and with it off a renormalised partial grid
            # sorts like any other score. The row says so either way so no
            # surface has to look the fact up for itself.
            "provisional": bool(
                (sub.get("evaluation") or {}).get("score_provisional")
                or (sub.get("evaluation") or {}).get("grid_complete") is False),
        }
        # The one field on this row that is a decision rather than a fact
        # about the candidate. Present only when the send was asked for it, so
        # every surface downstream can read "is there a score here?" off the
        # rows themselves instead of re-deriving the policy.
        if scores_on(include_scores):
            row["score"] = (sub.get("evaluation") or {}).get("score")
        out.append(row)
    return out


def held_back(job_id: int, tier: str = "") -> list[dict]:
    """
    Candidates kept off the shortlist because their grid came back part-filled.

    The recruiter's half of the partial-grid rule. Holding someone out of a
    ranking they do not belong in is right; doing it silently would swap one
    invisible failure for another, and the person who can fix it -- by
    re-grading -- is exactly the person looking at this screen.

    Returns the same row shape as `rows()`, without a rank: they are not
    ranked, that is the point. Empty is the normal case.
    """
    default_tier = rubric_pack.default_tier_for_slug(
        (store.get_role(job_id) or {}).get("slug")) if tier else None
    out = []
    for sub in store.held_back(job_id, tier=tier or None,
                               default_tier=default_tier):
        ev = sub.get("evaluation") or {}
        out.append({
            "submission_id": sub["_id"],
            "name": sub.get("candidate_name") or "(no name)",
            "email": sub.get("candidate_email") or "",
            "assessment_url": sub.get("admin_url") or "",
            "submitted_at": _fmt_date(sub.get("submitted_at")),
            # What was actually judged, so "re-grade this one" is a decision
            # with a number behind it rather than a shrug. `grid_of` is absent
            # on verdicts written before the coverage fields existed.
            "grid_marked": ev.get("grid_marked"),
            "grid_of": ev.get("grid_of"),
            "grid_coverage": ev.get("grid_coverage"),
            "unmarked": ev.get("grid_unmarked") or [],
        })
    return out


def _slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", str(text or "role").lower()).strip("-") or "role"


def filename(role: dict, extension: str = "xlsx") -> str:
    return f"shortlist-{_slug(role.get('slug') or role.get('title'))}.{extension}"


def dashboard_link(job_id: int) -> str:
    """
    The dashboard, opened on this role's shortlist tab.

    FOR RECRUITERS, NOT FOR THE MAIL. This URL is not put in front of a hiring
    manager and build_email() does not render it -- see the note there. It is
    returned to the dashboard's own frontend after a send, so the page can say
    where the list it just mailed lives.

    A deep link rather than the dashboard's front door. `#role=<id>&tab=
    shortlist` is the hash evaluations.js already reads on load and rewrites on
    every tab change, so this is the same URL a recruiter would get by copying
    their address bar while looking at the list -- not a second route that has
    to be kept in step with the page.

    Empty when the link is switched off, so a caller can pass the result
    straight through and the template's "nothing in, nothing out" rule does the
    rest.
    """
    if not SHORTLIST_DASHBOARD_LINK:
        return ""
    return f"{DASHBOARD_BASE_URL}/evaluations.html#role={job_id}&tab=shortlist"


def review_link(token: str) -> str:
    """The manager's review page URL for a token."""
    return f"{PUBLIC_BASE_URL}/review/{token}"


def is_loopback(base_url: str) -> bool:
    """
    True when PUBLIC_BASE_URL points at this machine.

    Worth checking before a send rather than after: a link built from
    127.0.0.1 opens fine for whoever clicked Send and is dead for every manager
    it was mailed to, and nothing about the delivered message looks wrong. The
    default is loopback, so this fires until somebody sets the value.
    """
    host = str(base_url or "").split("://")[-1].split("/")[0].split(":")[0].lower()
    return host in ("127.0.0.1", "localhost", "::1", "0.0.0.0", "")


# ---------------------------------------------------------------------------
# Spreadsheet
# ---------------------------------------------------------------------------

# Column key, header, width. One list drives the header row, the widths and the
# cell order, so a column added here cannot land in the sheet unlabelled.
COLUMNS = [
    ("rank", "#", 5),
    ("name", "Candidate", 26),
    ("email", "Email", 30),
    ("resume_link", "Resume", 14),
    ("assessment_url", "Assessment", 14),
    ("video_link", "Video", 14),
    ("submitted_at", "Submitted", 13),
]

# Columns whose value is a URL: shown as a word, linked to the address.
LINK_LABEL = {"resume_link": "Open CV", "assessment_url": "View answers",
              "video_link": "Watch"}


def build_xlsx(role: dict, shortlist: list[dict]) -> bytes:
    """
    The attached spreadsheet.

    Links are written as real hyperlinks behind a short label rather than as
    raw URLs: a Google Drive address is 90 characters of noise in a column the
    manager only ever clicks.

    THE SCORE COLUMN IS DECIDED BY THE ROWS, NOT BY THIS FUNCTION. It appears
    when the rows carry a `score`, which `rows(include_scores=...)` put there.
    Reading the flag a second time here is how the attachment and the download
    end up disagreeing about a send that asked for one and not the other, and
    the manager is the only person who would ever see both.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:                          # pragma: no cover
        raise ShortlistError(
            "openpyxl is not installed, so the spreadsheet cannot be built. "
            "Run: pip install openpyxl"
        ) from exc

    columns = list(COLUMNS)
    show_scores = any("score" in row for row in shortlist)
    if show_scores:
        columns.insert(2, ("score", "AI score", 10))
    # Only when there is something to say. A column of blanks on every normal
    # send would train people to ignore the one send where it is filled in.
    if any(row.get("provisional") for row in shortlist):
        columns.append(("grading", "Grading", 16))

    book = Workbook()
    sheet = book.active
    # Excel refuses a sheet name over 31 characters or containing []:*?/\.
    sheet.title = re.sub(r"[\[\]:*?/\\]", "", str(role.get("title") or "Shortlist"))[:31]

    title_font = Font(name="Calibri", size=14, bold=True, color="FF001D6B")
    sheet["A1"] = f"{role.get('title') or 'Role'} — top {len(shortlist)} candidates"
    sheet["A1"].font = title_font
    caption = ("Ranked by assessment review, strongest first. "
               "Click a link to open the candidate's CV or their answers.")
    if show_scores:
        # The number needs its sentence in the same file as the number. A
        # column headed "AI score" with nothing beside it gets read as a mark
        # somebody gave the person, and 78 vs 74 gets treated as a gap.
        caption += (" The AI score is our grader's mark out of 100 against this "
                    "role's rubric — a starting point for your own read of the "
                    "work, not a verdict on the candidate.")
    sheet["A2"] = caption
    sheet["A2"].font = Font(name="Calibri", size=10, italic=True, color="FF5B6270")

    head_row = 4
    header_fill = PatternFill("solid", fgColor="FF001D6B")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    for index, (_, label, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=head_row, column=index, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = width

    link_font = Font(name="Calibri", size=11, color="FF0B2E8E", underline="single")
    for offset, row in enumerate(shortlist, start=1):
        for index, (key, _, _) in enumerate(columns, start=1):
            cell = sheet.cell(row=head_row + offset, column=index)
            url = row.get(key) if key in LINK_LABEL else None
            if url:
                cell.value = LINK_LABEL[key]
                cell.hyperlink = url
                cell.font = link_font
            elif key in LINK_LABEL:
                cell.value = "—"
                cell.font = Font(name="Calibri", size=11, color="FF98A2B3")
            elif key == "grading":
                cell.value = "Partly graded" if row.get("provisional") else "Complete"
            elif key == "score":
                # Written as a number so the column sorts and filters as one,
                # rounded because the extra decimals the scorer carries are
                # precision this figure does not have. A row with no verdict
                # cannot happen on a shortlist -- it is ranked BY the score --
                # but it says so rather than showing an empty cell that reads
                # like a zero.
                value = row.get("score")
                cell.value = (round(float(value), 1)
                              if isinstance(value, (int, float)) else "—")
                cell.number_format = "0.0"
            else:
                cell.value = row.get(key)
            if key in ("rank", "score"):
                cell.alignment = Alignment(horizontal="center")

    # Freeze the header and turn on filters, so a 20-row sheet still behaves
    # like a list rather than a picture of one.
    sheet.freeze_panes = sheet.cell(row=head_row + 1, column=1)
    if shortlist:
        last = get_column_letter(len(columns))
        sheet.auto_filter.ref = f"A{head_row}:{last}{head_row + len(shortlist)}"

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Email
# ---------------------------------------------------------------------------

def _link(url: str, label: str) -> str:
    if not url:
        return '<span style="color:#98a2b3;">—</span>'
    return (f'<a href="{url}" style="color:#0b2e8e;font-weight:600;'
            f'text-decoration:none;">{label}</a>')


def _partial_tag(row: dict) -> str:
    """
    A rank the AI did not fully earn, said in the table it appears in.

    Normally never rendered: a provisional verdict is held out of the ranking
    upstream. With SHORTLIST_REQUIRE_COMPLETE_GRID off it is rendered on every
    copy of the row, because the manager sees rank and nothing else -- there is
    no score on this page for a caveat to attach to, so it attaches to the name.
    """
    if not row.get("provisional"):
        return ""
    return (f'<span style="margin-left:8px;font-size:11px;font-weight:600;'
            f'color:{MUTED};border:1px dashed {MUTED};border-radius:3px;'
            f'padding:1px 6px;">partly graded</span>')


def build_email(role: dict, shortlist: list[dict], to_name: str = "",
                note: str = "", review_url: str = "",
                dashboard_url: str = "", preview: bool = False) -> dict:
    """
    Render the shortlist email without sending it.

    The dashboard previews through this function, so what a recruiter reads
    before clicking send is what the manager receives -- with one deliberate
    exception, `preview`, described below.

    `note` is the recruiter's own line, dropped in above the table. It is
    escaped, not templated -- it comes from a text box, and a stray angle
    bracket in it should reach the manager as an angle bracket rather than
    breaking the layout.

    `review_url` is this manager's private link to their live workspace for
    this role: the page where they read the candidates, pick the ones worth
    meeting, write the invitation and send it. It is the whole point of the
    message, so it is a button above the table rather than a line under it -- a
    manager who reads only the first screen should still know there is
    something to click.

    `preview` says this render is for the recruiter's own screen and no token
    exists yet, so the button is drawn as an inert sample instead of being
    dropped. It USED TO BE DROPPED, on the reasoning that a button going
    nowhere teaches people not to press it -- but the preview is the only look
    anyone gets at this email before it leaves, and a preview missing the one
    thing the message exists for reads as "the link is broken" to the person
    who would have to fix it. Sending is unaffected: send_shortlist() mints a
    token per manager first, so a real message never takes this path.

    With neither a `review_url` nor `preview`, the block is omitted entirely,
    which is what any other caller building a message without a link wants.

    `dashboard_url` is the second link: this role's list on the recruiting
    dashboard, rendered as a line of text under the button. It is the lesser of
    the two -- deciding happens behind `review_url` -- so it is deliberately
    not a second button.

    IT IS ONLY SAFE BECAUSE THE DASHBOARD HAS ACCOUNTS. It lists every
    candidate on every role with their address and their score, and answers
    routes that mail hundreds of people; what makes it mailable is that
    auth.AUTH_ENABLED gates it and a manager's reach is scoped per request to
    the roles they are listed on. Turning AUTH_ENABLED off turns this line into
    an open door, and nothing here checks -- if that switch ever moves, this
    call site has to move with it.

    Pass "" to omit the block entirely, which is what a caller wants when
    SHORTLIST_DASHBOARD_LINK is off.
    """
    title = role.get("title") or "the role"
    count = len(shortlist)
    subject = f"Top {count} candidates — {title}"

    body_rows = "".join(f"""
        <tr>
          <td style="padding:12px 8px;border-bottom:1px solid {LINE};
                     color:{MUTED};font-variant-numeric:tabular-nums;">{row['rank']}</td>
          <td style="padding:12px 8px;border-bottom:1px solid {LINE};">
            <div style="font-weight:600;color:{INK};">{_esc(row['name'])}{
              _partial_tag(row)}</div>
            <div style="font-size:13px;color:{MUTED};">{_esc(row['email'])}</div>
          </td>
          <td style="padding:12px 8px;border-bottom:1px solid {LINE};">
            {_link(row['resume_link'], 'CV')}</td>
          <td style="padding:12px 8px;border-bottom:1px solid {LINE};">
            {_link(row['assessment_url'], 'Answers')}</td>
          <td style="padding:12px 8px;border-bottom:1px solid {LINE};">
            {_link(row['video_link'], 'Video')}</td>
        </tr>""" for row in shortlist)

    note_html = (f"""
            <p style="margin:0 0 20px;padding:12px 16px;background:#f6f7fb;
                      border-left:3px solid {NAVY};white-space:pre-wrap;">{_esc(note)}</p>"""
                 if note.strip() else "")

    # The action. A private link, so it says so -- a manager who forwards this
    # to a colleague should know they are handing over the ability to decide,
    # not just to look.
    #
    # The sentence under the button names the three things the page now does,
    # in the order they are done in. Interviews START there and nowhere else,
    # and a manager who does not know that will read this table, reply "let us
    # meet 2, 4 and 9", and wait for a recruiter who is waiting for them.
    #
    # In a preview there is no token to point at, so the same button is drawn
    # inert with a line saying why -- see the note on `preview`. The banner is
    # part of the preview and not of the mail, and it is worded so that a
    # recruiter who screenshots this screen cannot pass it off as the real
    # message.
    review_button = f"""
        <p style="margin:0 0 10px;">
          <a href="{_esc(review_url)}"
             style="display:inline-block;background:{NAVY};color:#ffffff;
                    padding:12px 22px;border-radius:8px;font-weight:600;
                    text-decoration:none;">Open your shortlist</a>
        </p>""" if review_url else f"""
        <p style="margin:0 0 10px;">
          <span style="display:inline-block;background:{MUTED};color:#ffffff;
                       padding:12px 22px;border-radius:8px;font-weight:600;
                       opacity:0.65;">Open your shortlist</span>
          <span style="color:{MUTED};font-size:12px;padding-left:10px;">
            sample — not a live link
          </span>
        </p>
        <p style="margin:0 0 10px;color:{MUTED};font-size:12px;
                  font-style:italic;">
          Preview only. Each manager's copy carries their own private link
          here, minted when you send.
        </p>"""

    review_html = (f"""{review_button}
        <p style="margin:0 0 24px;color:{MUTED};font-size:13px;">
          Your live page for this role: narrow the list down, tick the people
          you want to meet, and send them an invitation over your own calendar
          — in your words, after you have read it. Hire and rejection are
          on the same page. The link is private to you, so please do not
          forward it.
        </p>"""
                   if (review_url or preview) else "")

    # The secondary link: this role's list on the recruiting dashboard. Text
    # under the button rather than a second button, because deciding happens
    # behind `review_url` and two equal-looking buttons would make a reader
    # choose before they know which one they wanted. Omitted when there is no
    # URL, so a switched-off SHORTLIST_DASHBOARD_LINK leaves no orphaned
    # sentence behind it.
    #
    # LOCAL TESTING ONLY WHILE DASHBOARD_BASE_URL IS LOOPBACK. The default
    # address resolves on the machine that sent the mail and nowhere else, so
    # this link is dead for any real manager -- send_shortlist() reports that
    # as `board_unreachable` rather than leaving it to be noticed by silence.
    dashboard_html = (f"""
        <p style="margin:0 0 24px;color:{MUTED};font-size:13px;">
          Already met some of them? The
          <a href="{_esc(dashboard_url)}"
             style="color:{NAVY};font-weight:600;">shortlist board</a>
          shows where everyone on this list stands. Sign in with your own
          account.
        </p>"""
                      if dashboard_url else "")

    html = f"""
    <div style="font-family:'Poppins',Arial,sans-serif;max-width:720px;margin:0 auto;
                padding:32px;color:{INK};font-size:15px;line-height:1.6;">

      {brevo_client.header_html(720)}

      <div style="padding:32px;border:1px solid {LINE};border-top:none;
                  border-radius:0 0 8px 8px;">

        <p style="margin:0 0 16px;">Hi {_esc(_first_name(to_name))},</p>

        <p style="margin:0 0 16px;">
          Here are the top <strong>{count}</strong> candidates for
          <strong>{_esc(title)}</strong>, ranked strongest first after review of
          their assessments.
        </p>
        {note_html}
        <p style="margin:0 0 20px;color:{MUTED};font-size:14px;">
          Each row links to the candidate's CV and to the answers they submitted,
          so you can form your own view. The full list is attached as a
          spreadsheet.
        </p>
        {review_html}
        {dashboard_html}

        <table style="width:100%;border-collapse:collapse;font-size:14px;
                      margin:0 0 24px;">
          <thead>
            <tr style="text-align:left;">
              <th style="padding:8px;border-bottom:2px solid {NAVY};color:{MUTED};
                         font-size:12px;text-transform:uppercase;
                         letter-spacing:0.04em;font-weight:600;">#</th>
              <th style="padding:8px;border-bottom:2px solid {NAVY};color:{MUTED};
                         font-size:12px;text-transform:uppercase;
                         letter-spacing:0.04em;font-weight:600;">Candidate</th>
              <th style="padding:8px;border-bottom:2px solid {NAVY};color:{MUTED};
                         font-size:12px;text-transform:uppercase;
                         letter-spacing:0.04em;font-weight:600;">Resume</th>
              <th style="padding:8px;border-bottom:2px solid {NAVY};color:{MUTED};
                         font-size:12px;text-transform:uppercase;
                         letter-spacing:0.04em;font-weight:600;">Assessment</th>
              <th style="padding:8px;border-bottom:2px solid {NAVY};color:{MUTED};
                         font-size:12px;text-transform:uppercase;
                         letter-spacing:0.04em;font-weight:600;">Video</th>
            </tr>
          </thead>
          <tbody>{body_rows}</tbody>
        </table>

        <p style="margin:0 0 24px;">
          No need to reply with names — whoever you invite from your page
          hears from you directly, and we will see it on our side.
        </p>

        <p style="margin:0;">
          Best,<br>
          {_esc(BREVO_SENDER_NAME)}<br>
          <span style="color:{MUTED};font-size:13px;">Ajaia Hiring Team</span>
        </p>
      </div>
    </div>
    """

    lines = [
        f"Hi {_first_name(to_name)},",
        "",
        f"Here are the top {count} candidates for {title}, ranked strongest "
        f"first after review of their assessments.",
    ]
    if note.strip():
        lines += ["", note.strip()]
    lines += ["", "The full list is attached as a spreadsheet.", ""]
    if review_url or preview:
        lines += ["Your live page for this role -- narrow the list down, tick the",
                  "people you want to meet, and send them an invitation over your",
                  "own calendar. Hire and rejection are on the same page. This",
                  "link is private to you, so please do not forward it:",
                  # The plain-text half of the inert button. A recruiter who
                  # reads the preview's text tab has to be told the same thing
                  # the HTML tab tells them, or the two disagree about whether
                  # this email has a link in it.
                  review_url or "[preview only -- each manager's copy carries "
                                "their own private link here, minted when you "
                                "send]",
                  ""]
    for row in shortlist:
        lines.append(f"{row['rank']}. {row['name']} <{row['email']}>"
                     + ("  [partly graded]" if row.get("provisional") else ""))
        if row["resume_link"]:
            lines.append(f"   CV:      {row['resume_link']}")
        if row["assessment_url"]:
            lines.append(f"   Answers: {row['assessment_url']}")
        if row["video_link"]:
            lines.append(f"   Video:   {row['video_link']}")
    lines += [
        "",
        "No need to reply with names -- whoever you invite from your page hears "
        "from you directly, and we will see it on our side.",
    ]
    # The same secondary link the HTML part carries, in the same place. A text
    # reader that got one link and not the other would be reading a different
    # email.
    if dashboard_url:
        lines += ["",
                  "Already met some of them? The shortlist board shows where "
                  "everyone on",
                  "this list stands. Sign in with your own account:",
                  dashboard_url]
    lines += [
        "",
        "Best,",
        BREVO_SENDER_NAME,
        "Ajaia Hiring Team",
    ]

    return {"subject": subject, "html": html, "text": "\n".join(lines)}


# ---------------------------------------------------------------------------
# Send
# ---------------------------------------------------------------------------

def send_shortlist(job_id: int, limit: int = SHORTLIST_SIZE,
                   recipients: list[dict] | None = None,
                   note: str = "", tier: str = "",
                   include_scores: bool | None = None) -> dict:
    """
    Build and send a role's shortlist to its hiring managers.

    `recipients` overrides the role's stored managers, which is what the
    dashboard's "send to me first" does. Each manager gets their own message
    rather than one message with everyone on it: the greeting is by first name,
    and a hiring manager should not be able to read the other managers'
    addresses off the header of a mail about candidates.

    Partial success is reported, not raised. Two of three sends landing is a
    materially different situation from none of them landing, and a recruiter
    about to click again needs to know which.

    `include_scores` puts the AI score in the attached spreadsheet -- the same
    tick the recruiter set on the dashboard before clicking send, carried down
    so the mail matches the preview they read. None keeps the config default.
    It reaches the attachment and nothing else: the message body, the review
    page and the board are all unchanged by it.

    Every manager gets their OWN review link, minted here. Not one link shared
    between them: the token is the credential, so a shared one could not be
    revoked for a manager who left without cutting off the colleague still
    working the role, and no board move could say which of them made it.
    """
    role = store.get_role(job_id)
    if role is None:
        raise ShortlistError(f"No role with job id {job_id}.")
    # Named for the posting rather than the assignment from here down, so the
    # subject line, the sheet tab and the attachment all say which of the two
    # seats this list is for. Everything that addresses the ROLE -- the manager
    # list, the dashboard link, the send record -- keeps using job_id.
    role = role_view(role, tier)

    managers = recipients if recipients is not None else store.get_role_managers(job_id)
    managers = [m for m in (managers or []) if (m.get("email") or "").strip()]
    if not managers:
        raise ShortlistError(
            f"{role.get('title')} has no hiring manager assigned. "
            f"Add one before sending."
        )

    shortlist = rows(job_id, limit, tier=tier, include_scores=include_scores)
    withheld = held_back(job_id, tier=tier)
    if not shortlist:
        # Distinguished from "nothing graded yet", because the fix is
        # different: these people WERE graded, on a rubric the AI did not
        # finish, and re-grading them is one click rather than a whole run.
        if withheld and SHORTLIST_REQUIRE_COMPLETE_GRID:
            raise ShortlistError(
                f"{role.get('title')} has {len(withheld)} scored candidate(s), "
                f"but the AI did not finish the rubric for any of them, so none "
                f"can be ranked. Re-grade them, then send."
            )
        raise ShortlistError(
            f"{role.get('title')} has no scored candidates yet, so there is "
            f"nothing to send. Grade its submissions first."
        )

    if withheld:
        log.warning(
            "[%s] %d candidate(s) held off the shortlist on a part-filled "
            "grid: %s", role.get("title"), len(withheld),
            ", ".join(r["name"] for r in withheld))

    # Built once and attached to every copy -- the sheet does not depend on who
    # is reading it, and rebuilding it per manager would just be slower.
    sheet = build_xlsx(role, shortlist)
    attachment = [(filename(role), sheet)]

    submission_ids = [r["submission_id"] for r in shortlist]

    # One URL for every copy: it addresses the role, not the reader, and the
    # loop below has nothing of its own to add to it.
    board_url = dashboard_link(job_id)

    sent, failed, links = [], [], []
    for manager in managers:
        # Minted before the send and left in place if the send fails. An
        # orphaned token grants nothing on its own -- nobody has the URL --
        # whereas minting it afterwards would mean a manager holding an email
        # whose button 404s, which is the failure they cannot work around.
        token = store.create_review_link(job_id, manager, submission_ids)
        review_url = review_link(token)

        email = build_email(role, shortlist, manager.get("name", ""), note,
                            review_url, board_url)
        try:
            brevo_client.send_email(
                to=[{"email": manager["email"],
                     "name": manager.get("name") or manager["email"]}],
                subject=email["subject"],
                html=email["html"],
                text=email["text"],
                attachments=attachment,
                # Replies go to the recruiter who sent it, not to a no-reply
                # box -- "can I see number 4 first?" is the point of the mail.
                reply_to=BREVO_SENDER_EMAIL,
            )
            sent.append(manager["email"])
            links.append({"email": manager["email"], "url": review_url})
        except brevo_client.BrevoError as exc:
            log.error("Shortlist to %s failed: %s", manager["email"], exc)
            # Revoked rather than left live: this link was never delivered, so
            # the only way anyone holds it is out of our own logs.
            store.revoke_review_link(token)
            failed.append({"email": manager["email"], "error": str(exc)})

    if sent:
        store.record_shortlist_send(job_id, sent, len(shortlist), submission_ids)

    return {
        "role": role.get("title"),
        "count": len(shortlist),
        # What the attachment actually carried, read back off the rows rather
        # than off the argument -- the caller's answer to "did they get the
        # score?" should be the sheet's answer, not the request's.
        "scores": any("score" in row for row in shortlist),
        "sent": sent,
        "failed": failed,
        "links": links,
        # Who this list left out, and why. Reported on the send rather than
        # only in the log: the mail has gone by the time anyone reads a log,
        # and the recruiter reading this response is the one who can re-grade
        # them and send again.
        "held_back": withheld,
        # Said back to the caller rather than only logged: a link built from a
        # loopback address works on the machine that sent it and nowhere else,
        # and twenty candidates have already gone out by the time anyone
        # notices by hand.
        "unreachable": is_loopback(PUBLIC_BASE_URL),
        # The same question asked of the OTHER address in the mail. They are
        # separate settings and only one has a default that works
        # (DASHBOARD_BASE_URL falls back to PUBLIC_BASE_URL), so a deployment
        # that set only PUBLIC_BASE_URL mails a working review button beside a
        # dead board link, and nothing about the delivered message looks wrong.
        # Empty board_url is not unreachable: a link that was switched off was
        # never going anywhere.
        "board_unreachable": bool(board_url) and is_loopback(DASHBOARD_BASE_URL),
        # Handed back so the dashboard can say where the link it just mailed
        # points, without rebuilding it from a config the browser cannot read.
        "dashboard_url": board_url,
    }
